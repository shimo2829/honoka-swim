import streamlit as st
import pandas as pd
import os
import re
import math
import base64
import requests
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from streamlit_echarts import st_echarts, JsCode 

# ---------------------------------------------------------
# ログイン（パスワード認証）
# ---------------------------------------------------------
PASSWORD = st.secrets["PASSWORD"]

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

# ---------------------------------------------------------
# ログイン前ヘッダー（固定）
# ---------------------------------------------------------
st.markdown(
    """
    <style>
        .fixed-header {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            background-color: white;
            padding: 20px 30px;
            border-bottom: 2px solid #ddd;
            z-index: 1000000;
        }
        .header-title {
            font-size: 28px;
            font-weight: 700;
            margin: 0;
            color: #000;
        }
    </style>

    <div class="fixed-header">
        <div class="header-title">HONOKA Swimming Records Dashboard</div>
    </div>
    """,
    unsafe_allow_html=True
)

# ヘッダー分の余白
st.markdown("<div style='margin-top:120px;'></div>", unsafe_allow_html=True)

# ---------------------------------------------------------
# パスワード入力
# ---------------------------------------------------------
if not st.session_state.authenticated:
    pw = st.text_input("パスワードを入力してください", type="password")

    if pw == PASSWORD:
        st.session_state.authenticated = True
        st.rerun()
    elif pw != "0128":
        st.error("パスワードが違います")

    st.stop()

# ---------------------------------------------------------
# ページ設定
# ---------------------------------------------------------
st.set_page_config(
    page_title="HONOKA Swimming Records",
    layout="wide"
)

# ---------------------------------------------------------
# GitHub secrets 読み込み
# ---------------------------------------------------------
GITHUB_TOKEN = st.secrets["GITHUB_TOKEN"]
GITHUB_REPO = st.secrets["GITHUB_REPO"]
GITHUB_FILE_PATH = st.secrets["GITHUB_FILE_PATH"]

# ---------------------------------------------------------
# GitHub から Excel を取得
# ---------------------------------------------------------
def download_excel_from_github(repo, file_path, token, local_path="temp.xlsx"):
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"
    headers = {"Authorization": f"token {token}"}

    res = requests.get(url, headers=headers)

    if res.status_code == 200:
        content = base64.b64decode(res.json()["content"])
        with open(local_path, "wb") as f:
            f.write(content)
        return local_path
    else:
        st.error("GitHub からファイルを取得できませんでした")
        return None

# ---------------------------------------------------------
# GitHub へ Excel をアップロード
# ---------------------------------------------------------
def update_excel_to_github(local_path, repo, file_path, token, commit_message="Update Excel"):
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"

    with open(local_path, "rb") as f:
        content = f.read()

    encoded = base64.b64encode(content).decode()

    res = requests.get(url, headers={"Authorization": f"token {token}"})
    sha = res.json().get("sha", None)

    data = {
        "message": commit_message,
        "content": encoded,
        "sha": sha
    }

    res = requests.put(url, json=data, headers={"Authorization": f"token {token}"})
    return res.status_code in [200, 201]

# ---------------------------------------------------------
# 他のシートを消さずに、指定シートだけ更新する関数
# ---------------------------------------------------------
def save_sheet_without_deleting_others(excel_path, sheet_name, df):
    wb = load_workbook(excel_path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(excel_path)

# ---------------------------------------------------------
# 列名を正規化
# ---------------------------------------------------------
def normalize_columns(df):
    new_cols = []
    for col in df.columns:
        c = str(col)
        c = c.replace(" ", "").replace("　", "")
        c = c.replace("ヒヅケ", "日付")
        new_cols.append(c)
    df.columns = new_cols
    return df

# ---------------------------------------------------------
# 競泳表記 → 秒
# ---------------------------------------------------------
def time_to_seconds(t):
    if t is None:
        return None

    if isinstance(t, pd.Timestamp):
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1e6

    if isinstance(t, (int, float)) and t > 30000:
        return None

    if isinstance(t, (int, float)):
        if 0 < t < 1:
            return t * 86400
        else:
            return float(t)

    s = str(t).strip()
    s = s.replace("：", ":")

    m = re.match(r"(\d+)'(\d+)[\"”]?(\d+)", s)
    if m:
        minutes = int(m.group(1))
        seconds = int(m.group(2))
        ms = int(m.group(3))
        return minutes * 60 + seconds + ms / 100

    if ":" in s:
        try:
            m, sec = s.split(":")
            return int(m) * 60 + float(sec)
        except:
            pass

    try:
        return float(s)
    except:
        return None

# ---------------------------------------------------------
# 秒 → 競泳表記
# ---------------------------------------------------------
def seconds_to_swim_format(sec):
    if sec is None or (isinstance(sec, float) and math.isnan(sec)):
        return "―"
    m = int(sec // 60)
    s = sec % 60
    return f"{m}'{s:05.2f}"

# ---------------------------------------------------------
# Excel 読み込み
# ---------------------------------------------------------
local_excel = download_excel_from_github(GITHUB_REPO, GITHUB_FILE_PATH, GITHUB_TOKEN)
if local_excel is None:
    st.stop()

# ---------------------------------------------------------
# ★ ページ上部の種目選択（session_state 連動・完全版）
# ---------------------------------------------------------
event_list = ["フリー", "バッタ", "ブレ", "バック", "メドレー"]

# ① rerun の最初に session_state を確定
if "selected_event" not in st.session_state:
    st.session_state["selected_event"] = "フリー"

event = st.session_state["selected_event"]

# ② selectbox（key を「event」ではなく「固定文字列」にする）
event = st.selectbox(
    "種目を選択してください",
    event_list,
    index=event_list.index(event),
    key="event_selector"   # ← ここを固定にするのが正解
)

# ③ 選んだ event を保存
st.session_state["selected_event"] = event

# ---------------------------------------------------------
# 種目カラー
# ---------------------------------------------------------
event_colors = {
    "フリー": "#1E90FF",
    "バッタ": "#FF8C00",
    "ブレ":   "#32CD32",
    "バック": "#8A2BE2",
    "メドレー": "#DC143C"
}
title_color = event_colors.get(event, "#000000")

# ---------------------------------------------------------
# Excel 読み込み（距離選択より前に必ず実行）
# ---------------------------------------------------------
sheet_name = event

data = pd.read_excel(local_excel, sheet_name=sheet_name)
data = data.iloc[:, :6]
data.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]
data = normalize_columns(data)

data["タイム"] = data["タイム"].apply(time_to_seconds)
data["距離"] = pd.to_numeric(data["距離"], errors="coerce")
data = data.dropna(subset=["距離"])
data["距離"] = data["距離"].astype(int)

# ---------------------------------------------------------
# 距離選択
# ---------------------------------------------------------
if event == "メドレー":
    distance_list = [200, 400]
elif event == "ブレ":
    distance_list = [50, 100]
else:
    distance_list = sorted(data["距離"].unique())

# event が変わったら距離もリセット
if "selected_distance" not in st.session_state or st.session_state.get("last_event") != event:
    st.session_state["selected_distance"] = distance_list[0]

distance = st.selectbox(
    "距離を選択してください",
    distance_list,
    key=f"distance_selector_{event}"
)

st.session_state["selected_distance"] = distance
st.session_state["last_event"] = event

# ---------------------------------------------------------
# 固定ヘッダー（色も event に連動）
# ---------------------------------------------------------
st.markdown(
    f"""
    <style>
        .stAppViewContainer {{
            padding-top: 120px !important;
        }}

        .fixed-header {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            background-color: white;
            padding: 20px 30px;
            border-bottom: 2px solid #ddd;
            z-index: 1000000;
        }}

        .header-title {{
            font-size: 28px;
            font-weight: 700;
            margin: 0;
            color: #000;
        }}

        .header-sub {{
            font-size: 20px;
            font-weight: 600;
            margin: 0;
            color: {title_color};
        }}

        /* スマホ最適化 */
        @media screen and (max-width: 600px) {{
            .fixed-header {{
                padding: 10px 15px !important;
            }}
            .header-title {{
                font-size: 20px !important;
            }}
            .header-sub {{
                font-size: 14px !important;
            }}
            .stAppViewContainer {{
                padding-top: 90px !important;
            }}
        }}
    </style>

    <div class="fixed-header">
        <div class="header-title">HONOKA Swimming Records Dashboard</div>
        <div class="header-sub">{event} {distance}m 記録推移</div>
    </div>
    """,
    unsafe_allow_html=True
)


# ---------------------------------------------------------
# 長水路／短水路
# ---------------------------------------------------------
course = st.selectbox("長水路／短水路を選択", ["全記録", "短水路", "長水路"], key="course_selector")

# ---------------------------------------------------------
# データ絞り込み
# ---------------------------------------------------------
if course == "全記録":
    filtered = data[data["距離"] == distance].sort_values("日付")
else:
    filtered = data[
        (data["距離"] == distance) &
        (data["長水路or短水路"] == course)
    ].sort_values("日付")

filtered = filtered[filtered["タイム"].notna()]

if filtered.empty:
    st.error(f"{event} の {distance}m（{course}）のデータがありません")
    st.stop()
# ---------------------------------------------------------
# グラフ用データ整形
# ---------------------------------------------------------
filtered["日付_学年"] = (
    filtered["日付"].dt.strftime("%Y-%m-%d") + "（" + filtered["学年"] + "）"
)

filtered["タイム_表示"] = filtered["タイム"].apply(seconds_to_swim_format)

x_data = filtered["日付_学年"].tolist()
y_data = filtered["タイム"].tolist()
y_label = filtered["タイム_表示"].tolist()

y_min_raw = min(y_data)
y_max_raw = max(y_data)

if event == "メドレー":
    y_min = math.floor(y_min_raw / 10) * 10
    y_max = math.ceil(y_max_raw / 10) * 10
    y_interval = 10
else:
    y_min = math.floor(y_min_raw / 2) * 2
    y_max = math.ceil(y_max_raw / 2) * 2
    y_interval = 2

# ---------------------------------------------------------
# series_data（1本の線、点ごとに色を変える）
# ---------------------------------------------------------
series_data = [
    {
        "value": y_data[i],
        "label": y_label[i],               # 点の上のタイム表示用（残す）
        "date_grade": x_data[i],           # 日付（学年）
        "place": filtered["会場"].iloc[i], # 会場
        "itemStyle": {
            "color": "#3366FF" if filtered["長水路or短水路"].iloc[i] == "長水路" else "#FF3333"
        }
    }
    for i in range(len(y_data))
]

# ---------------------------------------------------------
# ECharts options（1本の線に戻す）
# ---------------------------------------------------------
options = {
    "legend": {
        "top": 0,
        "left": "center",
        "data": ["長水路", "短水路"],
        "textStyle": {"color": "#000"}
    },
   "tooltip": {
    "trigger": "axis",
    "formatter": JsCode("""
        function (params) {
            const p = params[0].data;
            return p.date_grade + "<br>" + p.place;
        }
    """)
},

  "xAxis": {"type": "category", "data": x_data},
"yAxis": {
    "type": "value",
    "inverse": False,
    "min": y_min,
    "max": y_max,
    "interval": y_interval,
    "axisLabel": {
        "formatter": JsCode("""
            function (value) {
                const m = Math.floor(value / 60);
                const s = (value % 60).toFixed(2).padStart(5, '0');
                return m + "'" + s;
            }
        """)
    }
},
"dataZoom": [{"type": "inside"}, {"type": "slider"}],

    "series": [
        {
            "type": "line",
            "data": series_data,
            "smooth": False,
            "lineStyle": {"color": "gray", "width": 2},
            "label": {
                "show": True,
                "position": "top",
                "formatter": JsCode("function (p) { return p.data.label; }"),
                "fontSize": 12
            }
        }
    ]
}

# ---------------------------------------------------------
# グラフ描画
# ---------------------------------------------------------
st_echarts(options=options, height="500px")

# ---------------------------------------------------------
# 新しい記録を追加（折りたたみ）
# ---------------------------------------------------------
with st.expander("＋ 新しい記録を追加（クリックで開く）"):

    st.subheader("新しい記録を追加")

    with st.form("add_record_form"):

        new_event = st.selectbox(
            "種目を選択してください",
            event_list,
            key="new_event_selector"
        )

        # 全種目で共通の距離リスト
        new_distance_list = [50, 100, 200, 400]

        new_distance = st.selectbox("距離を選択してください", new_distance_list)

        new_date = st.date_input("日付")

        # 学年リスト（修正フォームと統一）
        grade_list = ["小1","小2","小3","小4","小5","小6","中1","中2","中3"]
        new_grade = st.selectbox("学年", grade_list)

        new_course = st.selectbox("長水路 or 短水路", ["長水路", "短水路"])
# -----------------------------
# タイム入力（分・秒・100分の1秒）
# -----------------------------
col1, col2, col3 = st.columns([1, 1, 1])

with col1:
    new_min = st.selectbox(
        "分",
        list(range(0, 10)),
        index=0,
        key="new_min"
    )

with col2:
    new_sec = st.selectbox(
        "秒",
        list(range(0, 60)),
        index=0,
        key="new_sec"
    )

with col3:
    new_ms = st.selectbox(
        "100分の1秒",
        list(range(0, 100)),
        index=0,
        format_func=lambda x: f"{x:02d}",
        key="new_ms"
    )

# 秒に変換
new_time_sec = new_min * 60 + new_sec + new_ms / 100


        new_place = st.text_input("会場", value="菰野スイミング")

        submitted = st.form_submit_button("追加する")

    if submitted:
        new_time_sec = time_to_seconds(new_time_str)

        if new_time_sec is None:
            st.error("タイムの形式が正しくありません")
        else:
            new_row = pd.DataFrame([{
                "日付": pd.to_datetime(new_date),
                "学年": new_grade,
                "距離": int(new_distance),
                "長水路or短水路": new_course,
                "タイム": new_time_sec,
                "会場": new_place
            }])

            try:
                book = pd.read_excel(local_excel, sheet_name=new_event)
                book = normalize_columns(book)
                book = book.iloc[:, :6]
                book.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]

                updated = pd.concat([book, new_row], ignore_index=True)

                save_sheet_without_deleting_others(local_excel, new_event, updated)

                update_excel_to_github(
                    local_path=local_excel,
                    repo=GITHUB_REPO,
                    file_path=GITHUB_FILE_PATH,
                    token=GITHUB_TOKEN,
                    commit_message=f"Add record: {new_event} {new_distance}m"
                )

                st.session_state["selected_event"] = new_event

                st.success("記録を追加しました！（GitHub にも反映済み）")
                st.rerun()

            except Exception as e:
                st.error(f"Excel 書き込みエラー: {e}")

# ---------------------------------------------------------
# 記録の修正・削除（折りたたみ）
# ---------------------------------------------------------
with st.expander("＋ 記録の修正・削除（クリックで開く）"):

    st.subheader("記録の修正・削除")

    edit_df = filtered.copy().reset_index(drop=True)
    edit_df["行番号"] = edit_df.index

    st.dataframe(edit_df[["行番号", "日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]])

    target_index = st.number_input(
        "修正・削除する行番号を入力",
        min_value=0,
        max_value=len(edit_df)-1,
        step=1
    )

    target_row = edit_df.iloc[target_index]

    st.write("選択中の記録：")
    st.write(target_row)

    # -------------------------
    # 修正フォーム
    # -------------------------
    with st.form("edit_form"):

        grade_list = ["小1","小2","小3","小4","小5","小6","中1","中2","中3"]

        e_date = st.date_input("日付（修正）", value=target_row["日付"])
        e_grade = st.selectbox(
            "学年（修正）",
            grade_list,
            index=grade_list.index(target_row["学年"])
        )
        e_distance = st.number_input("距離（修正）", value=int(target_row["距離"]))
        e_course = st.selectbox(
            "長水路 or 短水路（修正）",
            ["長水路", "短水路"],
            index=0 if target_row["長水路or短水路"] == "長水路" else 1
        )
        e_time_str = st.text_input("タイム（修正）", value=seconds_to_swim_format(target_row["タイム"]))
        e_place = st.text_input("会場（修正）", value=target_row["会場"])

        edit_submitted = st.form_submit_button("修正する")

    # -------------------------
    # 修正処理
    # -------------------------
    if edit_submitted:

        new_time_sec = time_to_seconds(e_time_str)

        if new_time_sec is None:
            st.error("タイムの形式が正しくありません")
        else:
            book = pd.read_excel(local_excel, sheet_name=sheet_name)
            book = normalize_columns(book)
            book = book.iloc[:, :6]
            book.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]

            real_index = filtered.index[target_index]

            book.loc[real_index] = [
                pd.to_datetime(e_date),
                e_grade,
                int(e_distance),
                e_course,
                new_time_sec,
                e_place
            ]

            save_sheet_without_deleting_others(local_excel, sheet_name, book)

            update_excel_to_github(
                local_path=local_excel,
                repo=GITHUB_REPO,
                file_path=GITHUB_FILE_PATH,
                token=GITHUB_TOKEN,
                commit_message=f"Edit record: {event} {distance}m"
            )

            st.success("修正しました！（GitHub にも反映済み）")
            st.rerun()

    # -------------------------
    # 削除ボタン
    # -------------------------
    if st.button("この記録を削除する", type="primary"):

        try:
            book = pd.read_excel(local_excel, sheet_name=sheet_name)
            book = normalize_columns(book)
            book = book.iloc[:, :6]
            book.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]

            real_index = filtered.index[target_index]

            # 行削除
            book = book.drop(real_index).reset_index(drop=True)

            save_sheet_without_deleting_others(local_excel, sheet_name, book)

            update_excel_to_github(
                local_path=local_excel,
                repo=GITHUB_REPO,
                file_path=GITHUB_FILE_PATH,
                token=GITHUB_TOKEN,
                commit_message=f"Delete record: {event} {distance}m"
            )

            st.success("削除しました！（GitHub にも反映済み）")
            st.rerun()

        except Exception as e:
            st.error(f"削除中にエラーが発生しました: {e}")
